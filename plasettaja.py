import random
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill


df = pd.read_csv("osallistujat.csv")

poydat = {
    "Pöytä 1": [],
    "Pöytä 2": []
}

def jaa_istumapaikat(df, poydat):
    vieraat = df.sample(frac=1).reset_index(drop=True)
    
    for i, vieras in vieraat.iterrows():
        if i < 60:
            poydat["Pöytä 1"].append(vieras)
        else:
            poydat["Pöytä 2"].append(vieras)
    return poydat

def tallena_istumapaikat_excel(poydat, tiedostonimi="plase.xlsx"):
    wb = Workbook()
    
    ws1 = wb.active
    ws1.title = "Holilliset vs Holittomat"
    
    ws1.append(["Nimi", "Email", "Seuruetoive", "Ruokatoive", "Juomatoive", "Ainejärjestö", "Ilmoittautumisaika"])
    
    alkoholiton_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    alkoholillinen_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    
    for poyta, vieraat in poydat.items():
        for vieras in vieraat:
            row = [
                poyta,
                vieras['Nimi'],
                vieras['Email'],
                vieras['Seuruetoive'],
                vieras['Ruokatoive'],
                vieras['Juomatoive'],
                vieras['Ainejärjestö'],
                vieras['Ilmoittautumisaika'],
            ]
            
            ws1.append(row)
            
            if "alkoholiton" in vieras['Juomatoive'].lower():
                for col in range(1,9):
                    ws1.cell(row=ws1.max_row, column=col).fill = alkoholiton_fill
            elif "alkoholillinen" in vieras['Juomatoive'].lower():
                for col in range(1,9):
                    ws1.cell(row=ws1.max_row, column=col).fill = alkoholillinen_fill
                    
    ws2 = wb.create_sheet("Ruokavalion mukaan")
    ws2.append(["Nimi", "Email", "Seuruetoive", "Ruokatoive", "Juomatoive", "Ainejärjestö", "Ilmoittautumisaika"])


    gluteeniton_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    vegaani_fill = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")

    for poyta, vieraat in poydat.items():
        for vieras in vieraat:
            row = [
                poyta,
                vieras['Nimi'],
                vieras['Email'],
                vieras['Seuruetoive'],
                vieras['Ruokatoive'],
                vieras['Juomatoive'],
                vieras['Ainejärjestö'],
                vieras['Ilmoittautumisaika']
            ]
            ws2.append(row)
        
            if "gluteeniton" in vieras['Ruokatoive'].lower():
                for col in range(1,9):
                    ws2.cell(row=ws2.max_row, column=col).fill = gluteeniton_fill
            elif "vegaaninen" in vieras['Ruokatoive'].lower():
                for col in range(1,9):
                    ws2.cell(row=ws2.max_row, column=col).fill = vegaani_fill
                
    wb.save(tiedostonimi)

poydat = jaa_istumapaikat(df, poydat)
tallena_istumapaikat_excel(poydat)